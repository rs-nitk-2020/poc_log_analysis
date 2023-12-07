import pandas as pd
import openpyxl as px
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from typing import OrderedDict

class ReportOps:
    """
    Contains methods to generate formatted reports.
    """

    def generate_excel_report_data(report_path: str, df_data_dict: dict) -> dict:
        """
        Accepts a report path and a dictionary object that contains sheet names and dataframe values for sheet content
        and converts this into a formatted excel report that is stored in specified report path as well as returned as
        a dictionary.

        Args:
            report_path (str): path of the report 
            df_data_dict (dict): dictionary object that contains sheet names and dataframe values for sheet content

        Returns:
            dict: dictionary containing report data
        """
        
        report_data = dict()
        for sheet_name, dfs in df_data_dict.items():
            if not isinstance(dfs, (list, tuple)):
                df_data_dict[sheet_name] = [dfs]  

        # Write data to Excel file
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            for sheet_name, df_sheet in df_data_dict.items():
                # Merge dataframes if more than 1 log for same driver is present
                df_sheet = pd.concat(df_sheet, ignore_index=True)
                # Write to Excel file
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

        # Open the file using openpyxl to set visibility and apply styling
        workbook = px.load_workbook(report_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Apply borders to all cells with content
            cell_border = Border(left=Side(style='thick'), 
                                 right=Side(style='thick'), 
                                 top=Side(style='thick'), 
                                 bottom=Side(style='thick'))

            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                sheet.row_dimensions[row[0].row].height = 50
                for cell in row:
                    cell.border = cell_border
                    try:
                        numeric_value = float(cell.value)  # Try converting cell value to a number
                        if isinstance(numeric_value, (int, float)):
                            cell.alignment = Alignment(horizontal='center')  # Center align if the cell contains a number
                    except (ValueError, TypeError):
                        cell.alignment = Alignment(wrap_text=True, shrink_to_fit=True, horizontal='justify')  # Enable text wrapping and shrink to fit
                        pass # do nothing just checking if cell can be converted

            # Equalize column lengths
            column_lengths = [max(len(str(cell.value)) for cell in column) for column in sheet.iter_cols()]
            for col_idx, max_length in enumerate(column_lengths, start=1):
                col_letter = px.utils.get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].hidden = False
                sheet.column_dimensions[col_letter].width = 30  # Set width + 2 for padding
            
            # Apply styles to header cells (e.g., bold and background color)
            header_fill = PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type="solid")
            header_font = Font(bold=True)
            for cell in sheet[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align the content
                cell.fill = header_fill
                cell.font = header_font

            
        # Save the modified workbook to apply changes
        workbook.save(report_path)

        for sheet_name, df_list in df_data_dict.items():
            # Merge list of DataFrames into a single DataFrame
            merged_df = pd.concat(df_list, ignore_index=True)

            # Convert the merged DataFrame into an OrderedDict
            ordered_dict = OrderedDict(merged_df.to_dict(orient='list'))
            report_data[sheet_name] = ordered_dict

        return report_data
