import os
import re
import pandas as pd
# import dotenv
import openpyxl
from log_analysis_api.src.utils.report_utils import ReportOps as reports

# dotenv.load_dotenv()
log_directory = ""

df_data_dict = {}

driver_names = set()
driver_wise_file_names = {}
file_wise_counts = {}
driver_wise_counts = {}

log_sheet_header = ["Line No", "File Path", "Error No", "Error Message", "Misra Level", "MISRA Rule No"]
output_file_name = 'MISRA_Warning_Report.xlsx'


def parse_line(line, driver_name):
    parts = line.split()
    if len(parts) < 4:
        return None

    file_path = parts[0]
    pattern_hal = re.compile(r'hal/{}'.format(re.escape(driver_name)))
    pattern_mcal = re.compile(r'hal/mcal/{}'.format(re.escape(driver_name)))

    if pattern_hal.search(file_path):
        sheet_name = 'HAL_Log'
    elif pattern_mcal.search(file_path):
        sheet_name = 'MCAL_Log'
    else:
        return None  # Neither pattern is found in file path

    line_number = parts[1]
    error_no = parts[3].rstrip(":")

    driver_file_name = file_path.split('/')[-1].strip()

    error_message = ' '.join(parts[4:])
    error_message_part = error_message.split('[MM')[0].strip()

    if ' ' in error_message_part:
        error_message_part = f'"{error_message_part}"'

    misra_level_match = re.search(r'\bMM-PWT (\d+)', error_message)
    misra_level = misra_level_match.group(0) if misra_level_match else 'N/A'

    misra_rule_match = re.search(r'MISRA\s(.*?)(\d+\.\d+)', error_message)
    misra_rule_no = f"MISRA {misra_rule_match.group(1)} {misra_rule_match.group(2)}" if misra_rule_match else 'N/A'

    return [line_number, file_path, error_no, error_message_part, misra_level, misra_rule_no, driver_file_name, sheet_name]


def get_driver_name(filename):
    driver_name = filename.split("-")[0].lower()
    driver_names.add(driver_name)
    driver_wise_file_names.setdefault(driver_name, [])
    return driver_name


# prepare File_wise_calculation sheet
def prepare_file_wise_counts_sheet():
    print("Preparing file wise calculation sheet...")
    header = ['File Name', 'Driver', 'Count MM-PWT 0', 'Count MM-PWT 1',
              'Count MM-PWT 2', 'Count MM-PWT 3', 'Count MM-PWT 4',
              'Count MM-PWT 5', 'Count MM-PWT 6', 'Count MM-PWT 7',
              'File Wise Sum']
    sheet_name = 'File_wise_calculation'
    create_sheet(output_file_name, sheet_name, header)

    workbook = openpyxl.load_workbook(output_file_name)
    sheet = workbook[sheet_name]

    for driver_name in sorted(driver_names):
        for file_name in sorted(driver_wise_file_names[driver_name]):
            file_data = file_wise_counts.get(file_name, {})
            counts = [file_data.get(f"MM-PWT {i}", 0) for i in range(8)]
            sum_counts = sum(counts)
            row_data = [file_name, driver_name] + counts + [sum_counts]
            sheet.append(row_data)

    workbook.save(output_file_name)
    print("Done preparing file wise calculation sheet...")


def file_wise_counting(driver_name, driver_file_name, misra_level):
    if driver_name in driver_names and driver_file_name not in driver_wise_file_names.setdefault(driver_name, []):
        driver_wise_file_names[driver_name].append(driver_file_name)

    if driver_file_name not in file_wise_counts:
        file_wise_counts[driver_file_name] = {}

    file_wise_counts[driver_file_name][misra_level] = file_wise_counts[driver_file_name].get(misra_level, 0) + 1


# Final_output sheet
def prepare_driver_wise_counts_sheet():
    print("Preparing driver wise calculation sheet...")
    header = ['Driver Name', 'Driver Wise Sum', 'Count MM-PWT 0', 'Count MM-PWT 1',
              'Count MM-PWT 2', 'Count MM-PWT 3', 'Count MM-PWT 4',
              'Count MM-PWT 5', 'Count MM-PWT 6', 'Count MM-PWT 7' ]
    sheet_name = 'Final_output'
    create_sheet(output_file_name, sheet_name, header)

    workbook = openpyxl.load_workbook(output_file_name)
    sheet = workbook[sheet_name]

    for driver_name in sorted(driver_names):
        driver_counts = driver_wise_counts.get(driver_name, {})
        counts = [driver_counts.get(f"MM-PWT {i}", 0) for i in range(8)]
        sum_counts = sum(counts)
        row_data = [driver_name, sum_counts] + counts
        sheet.append(row_data)

    workbook.save(output_file_name)
    print("Done preparing driver wise calculation sheet...")


def driver_wise_counting():
    print("Started driver wise counting...")
    for driver_name in driver_names:
        driver_counts = {}
        for file_name in driver_wise_file_names.get(driver_name, []):
            file_counts = file_wise_counts.get(file_name, {})
            for i in range(8):
                misra_key = f"MM-PWT {i}"
                driver_counts[misra_key] = driver_counts.get(misra_key, 0) + file_counts.get(misra_key, 0)
        driver_wise_counts[driver_name] = driver_counts

    print("Done driver wise counting...")
    print(driver_wise_counts)


# HAL_Log & individual driver log files
def log_parsing():
    print("Started parsing log files...")
    print(log_directory)
    for root, _, files in os.walk(log_directory):
        for file in files:
            if not file.endswith('_LIN_LOG.txt'):
                continue
            driver_name = get_driver_name(file)
            file_path = os.path.join(root, file)

            create_sheet(output_file_name, driver_name, log_sheet_header)
            workbook = openpyxl.load_workbook(output_file_name)
            hal_sheet = workbook['HAL_Log']
            mcal_sheet = workbook['MCAL_Log']
            driver_sheet = workbook[driver_name]

            with open(file_path, 'r') as log_file:
                for line in log_file:
                    data = parse_line(line, driver_name)
                    if data == None:
                        continue

                    # [line_number, file_path, error_no, error_message_part, misra_level, misra_rule_no, driver_file_name, sheet_name]
                    _, _, _, _, misra_level, _, driver_file_name, sheet_name = data
                    file_wise_counting(driver_name, driver_file_name, misra_level)

                    data = data[:-2]
                    # write to driver_name sheet
                    driver_sheet.append(data)

                    if sheet_name == 'HAL_Log':
                        # write to HAL_Log sheet
                        hal_sheet.append(data)
                    elif sheet_name == 'MCAL_Log':
                        # write to MCAL_Log sheet
                        mcal_sheet.append(data)

            # save
            workbook.save(output_file_name)

    print("Done parsing log files...")


def create_sheet(file_name, sheet_name, header):
    workbook = openpyxl.load_workbook(file_name)
    driver_sheet = workbook.create_sheet(title=sheet_name)
    if header:
        driver_sheet.append(header)
    workbook.save(output_file_name)


def init_workbook():
    print("Initializing workbook...")
    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Rename the default sheet to 'HAL_Log'
    hal_sheet = workbook['Sheet']
    hal_sheet.title = 'HAL_Log'

    # Create a new sheet named 'MCAL_Log'
    mcal_sheet = workbook.create_sheet(title='MCAL_Log')

    # Append headers to both sheets
    hal_sheet.append(log_sheet_header)
    mcal_sheet.append(log_sheet_header)  # Assuming log_sheet_header is the same for both sheets

    # Save the workbook to a file
    workbook.save(output_file_name)
    print("Done creating workbook...")

def store_sheet_data(sheet_name):
    workbook = openpyxl.load_workbook(output_file_name)
    sheet = workbook[sheet_name]
    data = sheet.values
    cols = next(data)  # Extracting column names
    rows = list(data)  # Extracting data rows

    # Create a DataFrame using the extracted rows and columns
    df = pd.DataFrame(rows, columns=cols)

    # Store the DataFrame in df_data_dict using sheet_name as key
    df_data_dict[sheet_name] = df

def execute(log_directoryy, report_path, props):
    # print(log_directoryy)
    global log_directory
    log_directory = log_directoryy
    init_workbook()
    log_parsing()
    # print(driver_names)
    # print(driver_wise_file_names)
    # print(file_wise_counts)
    prepare_file_wise_counts_sheet()
    driver_wise_counting()
    prepare_driver_wise_counts_sheet()

    store_sheet_data('HAL_Log')
    store_sheet_data('MCAL_Log')
    for driver in driver_names:
        store_sheet_data(driver)
    store_sheet_data('File_wise_calculation')
    store_sheet_data('Final_output')
    # print(type(df_data_dict))
    # store_sheet_data('MCAL_file_wise_calculation')
    # store_sheet_data('MCAL_final_output')
    report_data = reports.generate_excel_report_data(report_path, df_data_dict)

    return report_data
